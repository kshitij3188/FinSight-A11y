#!/usr/bin/env python3
"""
Download all bunq.com/en-nl/documents into structured markdown files.
Handles:
  - HTML pages with full content
  - HTML shell pages that link to PDFs on static.bunq.com → follows and extracts
  - Direct PDF responses
  - XLSX files (Pillar 3 report)

Always overwrites existing files.

Run: python3 download_docs.py
Requires: pip install requests beautifulsoup4 html2text pdfplumber openpyxl
"""

import io
import os
import re
import time
import requests
from bs4 import BeautifulSoup
import html2text

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    import openpyxl
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False

BASE_URL = "https://www.bunq.com/en-nl/documents/"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bunq_legal")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

DOCUMENTS = [
    # ── HIGH PRIORITY ─────────────────────────────────────────────────────────
    ("terms",           "personal-account-terms-conditions"),
    ("terms",           "business-account-terms-conditions"),
    ("terms",           "api-terms-conditions"),
    ("terms",           "term-deposits-terms-conditions"),
    ("terms",           "crypto-terms-conditions"),
    ("terms",           "bunq-me-terms-conditions"),
    ("terms",           "mobile-wallet-terms-conditions"),
    ("terms",           "website-terms-conditions"),
    ("terms",           "idin-terms-conditions"),
    ("terms",           "bizum-terms-conditions"),
    ("privacy",         "privacy-policy"),
    ("privacy",         "data-processing-policy"),
    ("privacy",         "data-classification-policy"),
    ("privacy",         "data-retention-disposal-policy"),
    ("privacy",         "recruitment-privacy-policy"),
    ("fraud_complaints","fraud-refunds-policy"),
    ("fraud_complaints","recognizing-fraud-information-sheet"),
    ("fraud_complaints","suspecting-fraud-information-sheet"),
    ("fraud_complaints","formal-complaints-policy"),
    ("insurance",       "travel-insurance-terms-conditions"),
    ("insurance",       "travel-insurance-information-sheet"),
    ("insurance",       "purchase-protection-extended-warranty-terms-conditions"),
    ("insurance",       "metal-card-purchase-protection-extended-warranty-policy"),
    ("insurance",       "eligible-merchants-tap-to-pay"),
    ("regulatory",      "pricing-sheet"),
    ("regulatory",      "depositor-information-sheet"),
    ("regulatory",      "tax-statement-2025"),
    ("regulatory",      "accessibility-statement"),
    # ── MEDIUM PRIORITY ───────────────────────────────────────────────────────
    ("compliance",      "anti-money-laundering-counter-terrorist-financing-sanctions-policy"),
    ("compliance",      "anti-bribery-corruption-policy"),
    ("compliance",      "secure-banking-policy"),
    ("compliance",      "whistleblowing-policy"),
    ("compliance",      "code-of-conduct-policy"),
    ("compliance",      "conflicts-of-interest-policy"),
    ("compliance",      "corporate-governance-policy"),
    ("compliance",      "credit-risk-policy"),
    ("compliance",      "risk-management-policy"),
    ("compliance",      "third-party-risk-management-policy"),
    ("compliance",      "model-risk-management-policy"),
    ("compliance",      "operational-resilience-policy"),
    ("compliance",      "notification-of-incidents-policy"),
    ("compliance",      "responsible-disclosure-policy"),
    ("services",        "switch-service-terms-conditions"),
    ("services",        "bunq-esim-terms-conditions"),
    ("services",        "mastercard-chargeback-terms-conditions"),
    ("esg",             "environmental-social-governance-policy"),
    ("esg",             "socially-responsible-investing-policy"),
    ("esg",             "policy-on-esg-sri-criteria-for-borrowers-and-investments"),
    ("internal",        "dividend-policy"),
    ("internal",        "liquidity-policy"),
    ("internal",        "investment-strategy-information-sheet"),
    ("internal",        "procurement-policy"),
    ("internal",        "remuneration-policy"),
    # ── LOW PRIORITY ──────────────────────────────────────────────────────────
    ("promotions",      "rewards-terms-conditions"),
    ("promotions",      "bunq-points-terms-conditions"),
    ("promotions",      "bunq-cashback-partnerships-terms-conditions"),
    ("promotions",      "bunq-partners-terms-conditions"),
    ("promotions",      "bunq-x-rokt-cashback-terms-conditions"),
    ("promotions",      "bunq-switch-win-promotion-terms-conditions"),
    ("promotions",      "crypto-welcome-bonus-promotion-terms-conditions"),
    ("promotions",      "bunq-deals-general-terms-and-conditions"),
    ("promotions",      "bunq-deals-partnerships-terms-conditions"),
    ("promotions",      "bookingcom-cashback-terms-conditions"),
    ("promotions",      "bunq-x-lidl-cashback-terms-conditions"),
    ("promotions",      "bunq-x-auchan-cashback-terms-conditions"),
    ("promotions",      "bunq-x-albert-heijn-cashback-terms-conditions"),
    ("promotions",      "bunq-hyrox-instagram-giveaway-terms"),
    ("promotions",      "bunq-hyrox-raffle-promotion-terms-conditions"),
    ("reports",         "2024-annual-report"),
    ("reports",         "2024-pillar-3-report"),
    ("reports",         "2024-esg-report"),
    ("reports",         "2024-tax-report"),
    ("reports",         "2024-srep-decision-statement"),
    ("reports",         "2023-annual-report"),
    ("reports",         "2023-esg-report"),
    ("reports",         "2023-srep-decision-statement"),
    ("reports",         "2022-annual-report"),
    ("reports",         "2022-esg-report"),
    ("reports",         "2022-srep-decision-statement"),
    ("reports",         "2021-annual-report"),
    ("reports",         "2021-esg-report"),
    ("reports",         "2020-annual-report"),
    ("reports",         "2020-esg-report"),
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def slug_to_title(slug: str) -> str:
    return " ".join(w.capitalize() for w in slug.replace("-", " ").split())


def slug_to_filename(slug: str) -> str:
    return slug.replace("-", "_") + ".md"


def make_header(slug: str, source_url: str, actual_url: str = None) -> str:
    title = slug_to_title(slug)
    h = f"# {title}\n\n**Source:** {source_url}\n"
    if actual_url and actual_url != source_url:
        h += f"**Document URL:** {actual_url}\n"
    h += "\n---\n\n"
    return h


# ── Extractors ───────────────────────────────────────────────────────────────

def extract_pdf_bytes(raw: bytes, slug: str, source_url: str, doc_url: str) -> str:
    if not PDF_SUPPORT:
        return (
            make_header(slug, source_url, doc_url)
            + "_pdfplumber not installed. Run: pip install pdfplumber_\n"
        )
    try:
        pages = []
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for n, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text and text.strip():
                    pages.append(f"## Page {n}\n\n{text.strip()}")
        if pages:
            return make_header(slug, source_url, doc_url) + "\n\n".join(pages)
        return make_header(slug, source_url, doc_url) + "_No extractable text (scanned PDF?)_\n"
    except Exception as e:
        return make_header(slug, source_url, doc_url) + f"_PDF extraction error: {e}_\n"


def extract_xlsx_bytes(raw: bytes, slug: str, source_url: str, doc_url: str) -> str:
    if not XLSX_SUPPORT:
        return (
            make_header(slug, source_url, doc_url)
            + "_openpyxl not installed. Run: pip install openpyxl_\n"
        )
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sections = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append(" | ".join(cells))
            if rows:
                sections.append(f"## Sheet: {sheet_name}\n\n" + "\n".join(rows))
        if sections:
            return make_header(slug, source_url, doc_url) + "\n\n".join(sections)
        return make_header(slug, source_url, doc_url) + "_No data found in spreadsheet_\n"
    except Exception as e:
        return make_header(slug, source_url, doc_url) + f"_XLSX extraction error: {e}_\n"


def extract_html_content(raw: bytes, slug: str, source_url: str) -> tuple[str, list[str]]:
    """
    Parse HTML. Returns (markdown_text, list_of_pdf_urls_found).
    pdf_urls: any .pdf or .xlsx links pointing to static.bunq.com
    """
    html = raw.decode("utf-8", errors="replace")
    soup = BeautifulSoup(html, "html.parser")

    # Collect static file links before stripping
    file_links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "static.bunq.com" in href and (href.endswith(".pdf") or href.endswith(".xlsx")):
            file_links.append(href)

    # Remove noise
    for tag in soup.find_all(["nav", "header", "footer", "script", "style", "noscript", "iframe"]):
        tag.decompose()
    for tag in soup.find_all(class_=lambda c: c and any(
        k in " ".join(c if isinstance(c, list) else [c]).lower()
        for k in ["cookie", "banner", "overlay", "modal", "popup", "consent"]
    )):
        tag.decompose()

    main = (
        soup.find("main")
        or soup.find("article")
        or soup.find(attrs={"role": "main"})
        or soup.find(class_=lambda c: c and any(
            k in " ".join(c if isinstance(c, list) else [c]).lower()
            for k in ["document", "policy", "terms", "content", "legal"]
        ))
        or soup.find("body")
        or soup
    )

    h = html2text.HTML2Text()
    h.ignore_links = False
    h.ignore_images = True
    h.body_width = 0
    h.unicode_snob = True
    h.protect_links = True
    h.wrap_links = False

    md = h.handle(str(main)).strip()
    md = re.sub(r"\n{3,}", "\n\n", md)

    return md, file_links


def is_shell_page(md: str) -> bool:
    """True if the HTML page is just a 'this document is only available as PDF' shell."""
    text = md.lower()
    indicators = [
        "only available as a pdf",
        "only available as pdf",
        "open pdf",
        "download the",
    ]
    # Shell pages are short and contain PDF indicators
    return len(md) < 2000 and any(i in text for i in indicators)


# ── Fetcher ───────────────────────────────────────────────────────────────────

def fetch_bytes(url: str) -> tuple[bytes, str]:
    """Fetch URL, return (raw_bytes, content_type)."""
    resp = requests.get(url, headers=HEADERS, timeout=60, allow_redirects=True)
    resp.raise_for_status()
    return resp.content, resp.headers.get("content-type", "").lower()


def process_document(slug: str, source_url: str) -> tuple[str, str]:
    """
    Returns (markdown_content, status_label).
    status: html | pdf | xlsx | pdf_via_link | xlsx_via_link | empty | error
    """
    try:
        raw, ctype = fetch_bytes(source_url)

        # Direct PDF response
        if "pdf" in ctype or source_url.endswith(".pdf"):
            return extract_pdf_bytes(raw, slug, source_url, source_url), "pdf"

        # Direct XLSX response
        if "spreadsheet" in ctype or "excel" in ctype or source_url.endswith(".xlsx"):
            return extract_xlsx_bytes(raw, slug, source_url, source_url), "xlsx"

        # HTML — parse it
        md, file_links = extract_html_content(raw, slug, source_url)

        # If HTML is a shell page with a link to an actual file, follow it
        if file_links and is_shell_page(md):
            # Prefer PDF over xlsx, take first of each
            pdf_links  = [l for l in file_links if l.endswith(".pdf")]
            xlsx_links = [l for l in file_links if l.endswith(".xlsx")]

            results = []

            if pdf_links:
                doc_url = pdf_links[0]
                print(f"\n    → following PDF: {doc_url}", end="", flush=True)
                time.sleep(1)
                pdf_raw, _ = fetch_bytes(doc_url)
                results.append(extract_pdf_bytes(pdf_raw, slug, source_url, doc_url))

            if xlsx_links:
                doc_url = xlsx_links[0]
                print(f"\n    → following XLSX: {doc_url}", end="", flush=True)
                time.sleep(1)
                xlsx_raw, _ = fetch_bytes(doc_url)
                results.append(extract_xlsx_bytes(xlsx_raw, slug, source_url, doc_url))

            if results:
                combined = "\n\n---\n\n".join(results)
                return combined, "pdf_via_link" if pdf_links else "xlsx_via_link"

        # Plain HTML content
        if len(md.strip()) < 100:
            return make_header(slug, source_url) + "_No content found on this page._\n", "empty"

        return make_header(slug, source_url) + md, "html"

    except requests.exceptions.HTTPError as e:
        return (
            make_header(slug, source_url)
            + f"_HTTP {e.response.status_code} error fetching document._\n"
        ), "error"
    except Exception as e:
        return (
            make_header(slug, source_url)
            + f"_Fetch/parse error: {e}_\n"
        ), "error"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Dependencies:")
    print(f"  pdfplumber : {'✓' if PDF_SUPPORT  else '✗  pip install pdfplumber'}")
    print(f"  openpyxl   : {'✓' if XLSX_SUPPORT else '✗  pip install openpyxl'}")
    print()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for cat in set(c for c, _ in DOCUMENTS):
        os.makedirs(os.path.join(OUTPUT_DIR, cat), exist_ok=True)

    total   = len(DOCUMENTS)
    counts  = {}
    errors  = []

    print(f"Downloading {total} documents → {OUTPUT_DIR}/\n")

    for i, (category, slug) in enumerate(DOCUMENTS, 1):
        source_url = BASE_URL + slug
        filepath   = os.path.join(OUTPUT_DIR, category, slug_to_filename(slug))

        print(f"[{i:03d}/{total}] {slug} ... ", end="", flush=True)

        content, status = process_document(slug, source_url)

        counts[status] = counts.get(status, 0) + 1

        if status in ("error", "empty"):
            if os.path.exists(filepath):
                os.remove(filepath)
            if status == "error":
                print(f"FAILED (deleted if existed)")
                errors.append(slug)
            else:
                print(f"EMPTY (deleted if existed)")
        else:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(content)
            size_kb = len(content.encode()) / 1024
            print(f"OK [{status}] ({size_kb:.1f} KB)")

        time.sleep(1.5)

    print(f"\n{'='*60}")
    print(f"Total: {total}")
    for s, n in sorted(counts.items()):
        print(f"  {s:<20}: {n}")
    if errors:
        print(f"\nFailed:")
        for s in errors:
            print(f"  - {s}")
    print(f"\nFiles: {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
